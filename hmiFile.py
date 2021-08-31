# -*- coding: utf-8 -*-

import openpyxl
from math import ceil

def hmiFile(ws1,ws2,ws3,hmipath):
    #create HMITextLists-----------------------------------------------------
    #sheet1 textlist
    wb3=openpyxl.Workbook()
    ws5=wb3.active
    ws5.title='TextList'
    row=['Name','ListRange','Comment [zh-CN]','Comment [en-US]']
    ws5.append(row)
    row =['Cyl_Name','Cyl_Text_I0','Cyl_Text_I1','Cyl_Text_Q0','Cyl_Text_Q1','Wp_Name']
    for a in range(6):
        ws5.append([row[a],'Decimal','<No value>','<No value>'])
        
    #sheet2 TextListEntry
    ws6=wb3.create_sheet('TextListEntry')
    row=['Name','Parent','DefaultEntry','Value','Text [zh-CN]','Text [en-US]','FieldInfos']
    ws6.append(row)
    #cyl name----------------------------------------------------------------
    a=1
    for row in ws2.values:
        if row[0]:
            if str(row[0]).isdigit():
                b=int(ws2['L1'].value)*(int(row[0])-1)+int(row[1])
                ws6.append(['Text_list_entry_'+str(a),'Cyl_Name','',str(b),row[2],row[3]])
                a+=1
    #Cyl_Text_I0-------------------------------------------------------------
    a=1
    for row in ws2.values:
        if row[0]:
            if str(row[0]).isdigit():
                b=int(ws2['L1'].value)*(int(row[0])-1)+int(row[1])
                ws6.append(['Text_list_entry_'+str(a),'Cyl_Text_I0','',str(b),row[6],row[6]])
                a+=1
    #Cyl_Text_I1-------------------------------------------------------------
    a=1
    for row in ws2.values:
        if row[0]:
            if str(row[0]).isdigit():
                b=int(ws2['L1'].value)*(int(row[0])-1)+int(row[1])
                ws6.append(['Text_list_entry_'+str(a),'Cyl_Text_I1','',str(b),row[7],row[7]])
                a+=1  
    #Cyl_Text_Q0-------------------------------------------------------------
    a=1
    for row in ws2.values:
        if row[0]:
            if str(row[0]).isdigit():
                b=int(ws2['L1'].value)*(int(row[0])-1)+int(row[1])
                ws6.append(['Text_list_entry_'+str(a),'Cyl_Text_Q0','',str(b),row[4],row[4]])
                a+=1
    #Cyl_Text_Q1-------------------------------------------------------------
    a=1
    for row in ws2.values:
        if row[0]:
            if str(row[0]).isdigit():
                b=int(ws2['L1'].value)*(int(row[0])-1)+int(row[1])
                ws6.append(['Text_list_entry_'+str(a),'Cyl_Text_Q1','',str(b),row[5],row[5]])
                a+=1
    #wp name-----------------------------------------------------------------
    a=1
    for row in ws1.values:
        if row[0]:
            if str(row[0]).isdigit():
                ws6.append(['Text_list_entry_'+str(a),'WP_Name','',str(row[0]),row[1],row[2]])
                a+=1
                
    wb3.save(hmipath+"\\HMITextLists.xlsx")
    wb3.close()
    
    #create Cylinder Alarms-------------------------------------------------
    #create cylinder alarm CH----------------------------------------------
    wb3=openpyxl.Workbook()
    ws5=wb3.active
    ws5.title='DiscreteAlarms'
    
    row=['ID','Name','Alarm text [zh-CN], Alarm text','FieldInfo [Alarm text]','Class','Trigger tag','Trigger bit','Acknowledgement tag','Acknowledgement bit','PLC acknowledgement tag','PLC acknowledgement bit','Group','Report','Info text [zh-CN], Info text']
    ws5.append(row)
    a=1
    for row in ws2.values:
        if row[0]:
            if str(row[0]).isdigit():
                if not row[6]:
                    tempstr1=''
                else:
                    tempstr1=str(row[6])
                if not row[7]:
                    tempstr2=''
                else:
                    tempstr2=str(row[7])
                ws5.append([str(a),'Discrete_alarm_'+str(a),str(row[2])+' 原点报警 '+tempstr1,'','Warnings','Cylinder_DB_WorkPos{'+str(row[0])+'}_cylinderError{'+str(ceil(row[1]/8))+'}',str((row[1]*2-2)%16),'<No value>','0','<No value>','0','<No value>','True','<No value>'])
                ws5.append([str(a+1),'Discrete_alarm_'+str(a+1),str(row[2])+' 工作报警 '+tempstr2,'','Warnings','Cylinder_DB_WorkPos{'+str(row[0])+'}_cylinderError{'+str(ceil(row[1]/8))+'}',str((row[1]*2-1)%16),'<No value>','0','<No value>','0','<No value>','True','<No value>'])
                a+=2
    
    wb3.save(hmipath+"\\CylinderAlarms_CH.xlsx")
    wb3.close()
    #create cylinder alarm EN-----------------------------------------------
    wb3=openpyxl.Workbook()
    ws5=wb3.active
    ws5.title='DiscreteAlarms'
    
    row=['ID','Name','Alarm text [en-US], Alarm text','FieldInfo [Alarm text]','Class','Trigger tag','Trigger bit','Acknowledgement tag','Acknowledgement bit','PLC acknowledgement tag','PLC acknowledgement bit','Group','Report','Info text [en-US], Info text']
    ws5.append(row)
    a=1
    for row in ws2.values:
        if row[0]:
            if str(row[0]).isdigit():
                if not row[6]:
                    tempstr1=''
                else:
                    tempstr1=str(row[6])
                if not row[7]:
                    tempstr2=''
                else:
                    tempstr2=str(row[7])
                ws5.append([str(a),'Discrete_alarm_'+str(a),str(row[3])+' Home Alarm '+tempstr1,'','Warnings','Cylinder_DB_WorkPos{'+str(row[0])+'}_cylinderError{'+str(ceil(row[1]/8))+'}',str((row[1]*2-2)%16),'<No value>','0','<No value>','0','<No value>','True','<No value>'])
                ws5.append([str(a+1),'Discrete_alarm_'+str(a+1),str(row[3])+' Work Alarm '+tempstr2,'','Warnings','Cylinder_DB_WorkPos{'+str(row[0])+'}_cylinderError{'+str(ceil(row[1]/8))+'}',str((row[1]*2-1)%16),'<No value>','0','<No value>','0','<No value>','True','<No value>'])
                a+=2
    
    wb3.save(hmipath+"\\CylinderAlarms_EN.xlsx")
    wb3.close()
    
    #create HMI tages----------------------------------------------------
    wb3=openpyxl.Workbook()
    ws5=wb3.active
    ws5.title='Hmi Tags'
    
    row=['Name','Path','Connection','PLC tag','DataType','Length','Coding','Access Method','Address','Indirect addressing','Index tag','Start value','ID tag','Comment [zh-CN]','Acquisition mode','Acquisition cycle','Limit Upper 2 Type','Limit Upper 2','Limit Lower 2 Type','Limit Lower 2','Linear scaling','End value PLC','Start value PLC','End value HMI','Start value HMI']
    ws5.append(row)
    listCylError=[]
    for row in ws2.values:
        if row[0]:
            if str(row[0]).isdigit():
                a=[str(row[0]),str(ceil(row[1]/8))]
                if a in listCylError:
                    pass
                else:
                    listCylError.append(a)
    for each in listCylError:
        row=['Cylinder_DB_WorkPos{'+each[0]+'}_cylinderError{'+each[1]+'}','Cylinder_Alarm','HMI_连接_1','Cylinder_DB.WorkPos['+each[0]+'].cylinderError['+each[1]+']','Word','2','Binary','Symbolic access','<No Value>','False','<No Value>','<No Value>','0','<No Value>','Continuous','1 s','None','<No Value>','None','<No Value>','False','10','0','100','0']
        ws5.append(row)
    ws6=wb3.create_sheet('Multiplexing')
    row=['HMI Tag name','Multiplex Tag','Index']
    ws6.append(row)
    wb3.save(hmipath+"\\HMITags.xlsx")
    wb3.close() 